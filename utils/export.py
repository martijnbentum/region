from django.apps import apps
from utilities.models import instance2names
from django.core import serializers
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .model_util import compare_instances
from partial_date.partial_date import PartialDate

exclude_apps = 'admin,auth,contenttypes,django,sessions,utilities,easyaudit'.split(',')
non_primary_apps = ['locations']
all_models = apps.get_models()
selected_models = []
for m in all_models:
	app_name, model_name = instance2names(m)
	if app_name in exclude_apps or app_name in non_primary_apps: continue
	selected_models.append(m)
		
class Exports:
	def __init__(self, model_names= []):
		'''export all instances of all or a set of model names
		export_all 		export all models in the selected_models 
						this excludes models from the exclude apps and non_primary apps
						instance from the non_primary apps are still included if they are
						linked to included instances
		model_names 	export all instances from each model in model_names 
						instances of other models linked to these are also include
		'''
		self.export_all = True if not model_names else False
		if model_names:models = [m for m in all_models if instance2names(m)[1] in model_names]
		else: model_names = [instance2names(m)[1] for m in selected_models]
		self.model_names = model_names
		self.models = selected_models if self.export_all else models
		self.exports = []

	def make_exports(self):
		self.exports = []
		self._instances = []
		recursive = True if not self.export_all else False
		for i, model in enumerate(self.models):
			instances = model.objects.all()
			e = Export(instances,recursive)
			self.exports.append(e)
			self._instances.extend(e.instances)
		self._instances = list(set(self._instances))
		self.export = Export(self._instances,True)

	def to_excel(self,filename='default.xlsx',save = True):
		wb = self.export.to_excel(filename=filename,save=save)
		return wb

	@property
	def instances(self):
		return self.export.instances

	@property
	def xml(self):
		return self.export.xml

	@property
	def xml_str(self):
		return self.export.xml_str

	@property
	def json(self):
		return self.export.json
	
	@property
	def json_str(self):
		return self.export.json_str


class Export:
	'''export instances to different file formats.'''
	def __init__(self,instances, recursive = False):
		'''export instances to json, xml or excel
		instances 		list or queryset with instances to be exported
		recurive 		linked instances are automatically loaded, instance linked
						to the linked instances are also loaded when recursive is True
		'''
		self.recursive = recursive
		self.connection_set = ConnectionSet()
		self.connection_set.add_instances(instances, recursive)

	def __repr__(self):
		return 'export object for '+str(len(self.instances)) + ' instances' 

	def __str__(self):
		m = self.__repr__()
		m += 'XML representation:\n'
		m += self.xml_str[:1500] +'\n'
		m += '[...]'

	def to_excel(self,filename='default.xlsx', save =True):
		'''Stores the instance information in an excel workbook, each sheet contains
		instances of a specific model
		the sheet fieldtype holds all column names and field types for each model
		'''
		wb = Workbook()
		for e in self.xml.getchildren():
			o = XmlModelObject(e)
			wb = o.to_excel(wb)
		if save: wb.save(filename)
		return wb

	@property
	def instances(self):
		'''returns all instances (these also include instance related to the original set'''
		return self.connection_set.instances

	@property
	def xml(self):
		'''returns an lxml xml representation of all isntances.'''
		if not self.newly_added and hasattr(self,'_xml'):return self._xml
		x = serializers.serialize('xml',self.instances) 
		self._xml = etree.fromstring(bytes(x,encoding='utf-8'))
		return self._xml

	@property
	def xml_str(self):
		''' returns a str verson of the xml representation of the instances'''
		return etree.tostring(self.xml,pretty_print=True).decode()

	@property
	def json(self):
		''' returns a json representation of the instances'''
		self._json = json.loads(self.json_str)
		return self._json

	@property
	def json_str(self):
		''' returns a str version of the json representation of the instances'''
		if not self.newly_added and hasattr(self,'_json_str'):return self._json_str
		self._json_str = serializers.serialize('json',self.instances)
		return self._json_str

	@property
	def newly_added(self):
		'''flag to monitor if there are newly added instances
		if true  a new json xml etc is created.
		'''
		return self.connection_set.pk.newly_added
		

class Relations:
	'''finds all models that have a relation with this models.
	there are three types of relations:
		-a relation model (connecting this models with another models with potential other fields
		-a foreign key, a relation to another model
		-m2m, a relation to 1 one or more instances of another model
	these relation type are stored seperatly
	both the fields and model linked to those fields are stored 
	'''
	def __init__(self,model):
		self.app_name, self.model_name = instance2names(model)
		#retrieve all relation model ie model:Text -> TextPublication relation etc.
		self.relation_fields=[f for f in model._meta.get_fields() if f.one_to_many]
		self.relation_fields_str=[f.get_accessor_name() for f in self.relation_fields]
		self.relation_models = [f.related_model for f in self.relation_fields]
		#retrieve all foreign key fields and related models
		self.fk_fields = [f for f in model._meta.local_fields if f.is_relation]
		self.fk_fields_str= [f.get_cache_name() for f in self.fk_fields]
		self.fk_models = [f.related_model for f in self.fk_fields]
		#retrieve m2m fields and models
		self.m2m_fields = [f for f in model._meta.local_many_to_many if f.related_model != model]
		self.m2m_fields_str = [f.get_attname() for f in self.m2m_fields]
		self.m2m_models = [f.related_model for f in self.m2m_fields]

	def __repr__(self):
		m = 'relations of '
		m += 'app: '+self.app_name +' | '
		m += 'model: '+self.model_name 
		return m

	def __str__(self):
		m = 'app: '+self.app_name +'\n'
		m += 'model: '+self.model_name +'\n'
		m += 'relation models:\n\t' + model_list2str(self.relation_models,'\n\t') + '\n\n'
		m += 'fk models:\n\t' + model_list2str(self.fk_models,'\n\t') + '\n\n'
		m += 'm2m models:\n\t' + model_list2str(self.m2m_models, '\n\t')+'\n\n'
		return m


class ConnectionSet:
	'''collect all connections for one or more instances.
	can also recursively collect all connections for one or more instances
	i.e. find all connections for instances connected to the original set of instances
	'''
	def __init__(self):
		self.pk = Pk()
		self.connections = []
		self.input_instances = []
		self.counter = 0

	def __repr__(self):
		return 'connections for ' + str(len(self.input_instances)) + ' instances'

	def __str__(self):
		m = self.__repr__()
		return m + self.pk.__str__()

	def add_instances(self,instances, recursive = False):
		for instance in instances:
			if hasattr(instance,'endnode'):
				self.input_instances.append(instance)
				continue
			self.add_instance(instance)
		if recursive: self._run_recursive()

	def add_instance(self, instance, recursive = False):
		self.input_instances.append(instance)
		c = Connections(instance)
		self.pk.update(c.pk)
		self.connections.append(c)
		if recursive: self._run_recursive()

	def _run_recursive(self):
		self.counter += 1
		dif = list(set(self.instances) - set(self.input_instances))
		print('finding connections recursively, order:',self.counter,'instances:',len(dif))
		if dif: self.add_instances(dif,True)
		
	@property
	def instances(self):
		return self.pk.instances



class Connections:
	'''find all pk, app_name and model name of all connected instances of a given instance.'''
	def __init__(self,instance):
		self.instance = instance
		self.app_name, self.model_name = instance2names(instance)
		self.name = self.app_name + ' ' + self.model_name
		#retrieve all relation model ie model:Text -> TextPublication relation etc.
		self.relations = Relations(instance)
		self.pk = Pk()
		self.pk.add_instance(instance)
		for f in self.relations.relation_fields_str + self.relations.m2m_fields_str:
			rm = getattr(instance,f)
			self.pk.add_qeuryset(rm.all())
		for f in self.relations.fk_fields_str:
			ri = getattr(instance,f)
			self.pk.add_instance(ri)


	def __repr__(self):
		return 'connections for ' + self.name +' : '+ str(self.instance)

	def __str__(self):
		m = self.__repr__()
		return m + '\n' + str(self.pk)

	@property
	def instances(self):
		return self.pk.instances


class Pk:
	'''contain all pk for the different apps and models'''
	def __init__(self):
		self.d = {}
		self.d_newly_added = {}
		self.newly_added = True
		self._instances = []

	def add_instance(self,instance):
		if not instance:return
		k = instance2names(instance)
		if k not in self.d: self.d[k] = []
		if k not in self.d_newly_added: self.d_newly_added[k] = []
		if instance.pk not in self.d[k]: 
			self.newly_added = True
			self.d[k].append( instance.pk ) 
			self.d_newly_added[k].append( instance.pk)

	def add_qeuryset(self,qs):
		for instance in qs:
			self.add_instance(instance)

	def __repr__(self):
		lines = [', '.join(line) for line in self.d.keys()]
		values = [', '.join(map(str,self.d[key])) for key in self.d.keys()]
		longest = getlongeststr(lines) + 3
		m = '\n'.join([l.ljust(longest) + v for l, v in zip(lines,values)])
		return m

	def __str__(self):
		return self.__repr__()
		
	def update(self,other):
		for k in other.d:
			if k not in self.d: self.d[k] = []
			if k not in self.d_newly_added: self.d_newly_added[k] = []
			for value in other.d[k]:
				if value not in self.d[k]:
					self.newly_added = True
					self.d[k].append(value)
					self.d_newly_added[k].append(value)
				
	@property
	def instances(self):
		if self.newly_added:
			for k in self.d_newly_added.keys():
				model = apps.get_model(*k)
				for pk in self.d_newly_added[k]:
					self._instances.append(model.objects.get(pk =pk))
			self.newly_added = False
			self.d_newly_added = {}
		return self._instances
		
	
def model_list2str(ml,sep = '\t'):
	return sep.join([str(m) for m in ml]) 
		
def queryset2pk(qs):	
	return [i.pk for i in qs]
	
def getlongeststr(lines):
	longest = 0
	for line in lines:
		if len(line) > longest: longest = len(line)
	return longest

def format_dict(d, extra = 3):
	longest = getlongeststr(list(d.keys())) + extra
	return '\n'.join([key.ljust(longest) + val.__repr__() for key,val in d.items()])

class XmlModelObject:
	def __init__(self,xml):
		self.xml = xml
		self.app_name, self.model_name = xml.attrib['model'].split('.')
		self.pk = xml.attrib['pk']
		self.fields = [XmlFieldObject(f) for f in xml.getchildren()]
		self.column_names, self.column_values = ['pk'],[self.pk]
		self.column_type_names, self.column_type_values = [],[]
		for f in self.fields:
			self.column_names.append(f.column_name)
			self.column_values.append(f.column_value)
			self.column_type_names.extend(f.column_type_names)
			self.column_type_values.extend(f.column_type_values)


	def __repr__(self):
		return self.model_name + ' ' + str(self.pk)

	def __str__(self):
		m = self.__repr__() + '\n'
		m += '\n\n'.join([str(f) for f in self.fields])
		return m

	def to_excel(self,wb = None):
		if not wb: self.wb = Workbook()
		else: self.wb = wb
		if not self.model_name in self.wb.sheetnames: 
			self._set_fieldtypes()
			# self.type_sheet = self.wb.create_sheet(self.model_name+ '|type')
			self.sheet = self.wb.create_sheet(self.model_name)
			self._set_column_names()
			self.current_row = 2
		else:
			self.sheet = self.wb[self.model_name]
			# self.type_sheet = self.wb[self.model_name+'|type']
			self.current_row = self.sheet.max_row + 1
		self._set_values()
		if 'Sheet' in self.wb.sheetnames:
			self.wb.remove(self.wb['Sheet'])
		return self.wb

	def _set_column_names(self):
		for i, col in enumerate(self.column_names):
			self.sheet.cell(row=1,column=i+1,value=col)
			column = get_column_letter(i+1)
			self.sheet.column_dimensions[column].width = len(col)*1.2 +3
		'''
		for i, col in enumerate(self.column_type_names):
			self.type_sheet.cell(row=1,column=i+1,value=col)
			column = get_column_letter(i+1)
			self.type_sheet.column_dimensions[column].width = len(col) +3
		for i, val in enumerate(self.column_type_values):
			self.type_sheet.cell(row=2,column=i+1,value= val)
		'''

	def _set_fieldtypes(self):
		if not 'fieldtypes' in self.wb.sheetnames:
			self.fieldtype_sheet = self.wb.create_sheet('fieldtypes')
			for i, val in enumerate('app_name,model_name,fieldname,fieldtype,to'.split(',')):
				self.fieldtype_sheet.cell(row=1,column=i+1,value = val)
				column = get_column_letter(i+1)
				self.fieldtype_sheet.column_dimensions[column].width = 22
		else: self.fieldtype_sheet = self.wb['fieldtypes']
		self.current_type_row = self.fieldtype_sheet.max_row +1
		for f in self.fields:
			for i, val in enumerate([self.app_name,self.model_name]+f.field_type_values):
				self.fieldtype_sheet.cell(row=self.current_type_row,column=i+1,value =val)
			self.current_type_row +=1
			
		

	def _set_values(self):
		for i, val in enumerate(self.column_values):
			self.sheet.cell(row=self.current_row,column=i+1,value= val)
	

class XmlFieldObject:
	def __init__(self,xml):
		f = xml
		self.xml = xml
		self.name = f.attrib['name']
		self.column_name = self.name
		self.column_type_values,self.column_type_names = [], []
		if 'rel' in f.attrib:
			self.relational = True
			self.local = False
			self.type = f.attrib['rel']
			self.to = f.attrib['to'].split('.')[-1]
			self.column_type_names.append(self.name+'|to')
			self.column_type_values.append(self.to)
		if 'type' in f.attrib:
			self.relational = False
			self.local = True
			self.type = f.attrib['type']
			self.to = ''
		self.column_type_names.append(self.name+'|ftype')
		self.column_type_values.append(self.type)
		self.field_type_values = [self.name,self.type,self.to]

		if self.type == 'ManyToManyRel':
			self.value = ';'.join([c.attrib['pk'] for c in self.xml.getchildren()])
			self.column_value =self.value
		else: 
			self.value = str(xml.text) if xml.text else ''
			self.column_value =self.value


	def __repr__(self):
		m = 'relational field' if self.relational else 'local field, '
		m += 'name\t'+self.name 
		return m

	def __str__(self):
		m = 'relational field\n' if self.relational else 'local field\n'
		m += 'name\t'+self.name +'\n'
		m += 'type\t'+self.type+'\n'
		m += 'value\t'+self.value+'\n'
		if self.to: m += 'to\t' + self.to +'\n'
		return m
		
			


