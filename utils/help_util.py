#based on heritage of hunger help_util module
import json
from openpyxl import load_workbook


class Helper:
	def __init__(self, filename='data/Protocols.xlsx',model_name=None):
		self.filename = filename
		self.model_name = model_name
		self.wb = load_workbook(filename)
		self.model_helpers = {}
		self._set_values()

	def _set_values(self):
		for sheet_name in self.wb.sheetnames:
			if self.model_name != None and sheet_name.lower() != self.model_name.lower():continue
			self.model_helpers[sheet_name] = ModelHelper(self.wb,sheet_name)

	def get_dict(self,name=None):
		if self.model_name != None:name = self.model_name
		for key in self.model_helpers.keys():
			if name.lower() == key.lower(): 
				model_helper = self.model_helpers[key]
				break
		return model_helper._make_dict()

	def get_json(self,name=None):
		d = self.get_dict(name)
		return json.dump(d)
		

class ModelHelper:
	def __init__(self,wb,name):
		self.wb = wb
		self.name = name
		self.sheet = wb[name]
		self.field_helpers = {}
		self._set_values()

	def _set_values(self):
		self.values = list(self.sheet.values)
		self.errors = []
		for line in self.values:
			if len(line) < 2 or line[1] == 'Field':continue
			fh = FieldHelper(line)
			if fh.ok: self.field_helpers[fh.field_name] = fh
			else: self.errors.append(line)
				
	def _make_dict(self):
		d = {}
		for fh in self.field_helpers.values():
			d[fh.field_name] =fh.help_text
			d[fh.field_name.lower().strip()] =fh.help_text_html
		return d
			
		

class FieldHelper:
	def __init__(self,line):
		self.line = line
		self.ok = True
		try:
			self.model_name, self.field_name, self.help_text = line[:3]
			if self.field_name == None: self.ok = False
		except:self.ok = False
		if self.ok: 
			self.id = self.field_name.lower().replace(' ','_')
			self.help_text_html = ''
			for line in self.help_text.split('\n'):
				self.help_text_html += '<p>'+line+'</p> '
			

	def __repr__(self):
		return self.model_name + ' | ' + self.id

	def __str__(self):
		m = self.__repr__()
		m +='\n' + self.help_text
		return m
	

	
		
