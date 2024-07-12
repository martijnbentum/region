import Levenshtein
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import re

def load_locations():
    from locations.models import Location
    locations = Location.objects.all()
    return locations

def select_manually_added_locations(locations = None):
    if not locations: locations = load_locations()
    output = [x for x in locations if not re.match(r'^\d+$', x.geonameid)]
    return output

def check_location_matches(manually_added_locations = None, locations = None,
    max_distance = 2):
    if not locations: locations = load_locations()
    if not manually_added_locations: 
        manually_added_locations = select_manually_added_locations(locations)
    output = {}
    for location in manually_added_locations:
        name = location.name
        pk = location.pk
        for l in locations:
            if l.pk == pk: continue
            if check_match(name, l.name, max_distance):
                if not pk in output.keys(): output[pk] = []
                output[pk].append(l.pk)
    return output

def check_match(str1,str2, min_distance = 0, max_distance = 2):
    distance = Levenshtein.distance(str1, str2)
    match = min_distance <= distance <= max_distance
    return match


def load_texts():
    from catalogue.models import Text
    texts = Text.objects.all()
    return texts

def check_text_matches(texts = None, min_distance = 1, max_distance = 2):
    if not texts: texts = load_texts()
    output = {}
    for text1 in texts:
        setting = text1.setting
        if not setting: continue
        pk = text1.pk
        for text2 in texts:
            if text2.pk == pk: continue
            if not setting: continue
            if text2.pk in output.keys(): 
                if pk in output[text2.pk]: continue
            if check_match(setting, text2.setting, min_distance, max_distance):
                if not pk in output.keys(): output[pk] = []
                output[pk].append(text2.pk)
    return output

def pk_to_url_and_setting(pk):
    from catalogue.models import Text
    t = Text.objects.get(pk = pk)
    url = t.edit_url_complete
    setting = t.setting
    return [url, setting]

def make_workbook():
    return Workbook()

def add_sheet(wb, sheet_name):
    return wb.create_sheet(sheet_name)
    

def add_cell_value(sheet, row, column, value, hyperlink = ''):
    cell = sheet.cell(row = row, column = column, value = value)
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.style = 'Hyperlink'
    return cell

def to_output(almost_text_matches = None):
    if not almost_text_matches: almost_text_matches = check_text_matches()
    output = []
    for k,v in almost_text_matches.items():
        url1, setting1 = pk_to_url_and_setting(k)
        for pk in v:
            url2, setting2 = pk_to_url_and_setting(pk)
            output.append([k,pk,setting1,setting2,url1,url2])
    return output
    
def handle_line(line, row_index, sheet):
    pk1, pk2, setting1, setting2, url1, url2 = line
    add_cell_value(sheet, row_index+2, 1, pk1)
    add_cell_value(sheet, row_index+2, 2, setting1)
    add_cell_value(sheet, row_index+2, 3, setting2)
    add_cell_value(sheet, row_index+2, 4, 'link1', url1,)
    add_cell_value(sheet, row_index+2, 5, 'link2', url2,)
    add_cell_value(sheet, row_index+2, 6, pk1)
    add_cell_value(sheet, row_index+2, 7, pk2)

def make_text_setting_excel(output = None, filename = 'text_settings.xlsx'):
    if not output: output = to_output()
    wb = make_workbook()
    sheet = add_sheet(wb, 'items')
    for row_index, line in enumerate(output):
        handle_line(line, row_index, sheet)
    sheet.column_dimensions['B'].width = 100
    sheet.column_dimensions['C'].width = 100
    wb.remove(wb['Sheet'])
    wb.save(filename)
    return wb
    


            
            
            

            


